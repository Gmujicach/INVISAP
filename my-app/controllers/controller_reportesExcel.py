from flask import Blueprint, render_template, request, session, redirect, url_for, flash, send_file
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
import datetime
from models.model_reportesExcel import ReporteExcelModel
from services.bitacora_service import BitacoraService

reporte_excel_bp = Blueprint('reporte_excel_bp', __name__, template_folder='../vista')
modelo_reporte = ReporteExcelModel()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGO_GOBERNACION = os.path.join(BASE_DIR, 'static', 'assets', 'img', 'NUEVO LOGO GOBERNACION DE LARA JUNIO 2025.png')
LOGO_INVILARA = os.path.join(BASE_DIR, 'static', 'assets', 'img', 'INVILARA LOGO OFICIAL HORIZONTAL.png')

header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def construir_modulos_config_excel():
    return [
        {
            'id': 'solicitudes',
            'label': 'SOLICITUDES',
            'keywords': ['solicitudes', 'solicitud', 'solicitantes', 'solicitante'],
            'fetch': modelo_reporte.obtener_solicitudes_reporte,
            'headers': ['ID', 'Fecha', 'Tipo', 'Estatus', 'Problemática', 'Cédula', 'Prioridad'],
            'fields': ['id_solicitudes', 'fecha', 'tipo_solicitud', 'estatus_solicitud', 'problematica', 'cedula', 'prioridad']
        },
        {
            'id': 'empleados',
            'label': 'PERSONAL / EMPLEADOS',
            'keywords': ['empleados', 'empleado', 'personal', 'trabajadores', ' RRHH', 'recursos humanos'],
            'fetch': modelo_reporte.obtener_empleados_reporte,
            'headers': ['Nombre', 'Profesión', 'Gerencia', 'Ingreso', 'Email', 'Teléfono', 'Estado'],
            'fields': ['nombre_empleado', 'profesion_empleado', 'gerencia_asignada', 'fecha_ingreso', 'email_empleado', 'telefono_empleado', 'estado_empleado']
        },
        {
            'id': 'usuarios',
            'label': 'USUARIOS DEL SISTEMA',
            'keywords': ['usuarios', 'usuario', 'accesos', 'seguridad', 'roles'],
            'fetch': modelo_reporte.obtener_usuarios_reporte,
            'headers': ['Nombre', 'Cédula', 'Correo', 'Rol'],
            'fields': ['nombre', 'cedula_usuario', 'correo', 'rol']
        },
        {
            'id': 'contrataciones',
            'label': 'CONTRATACIONES',
            'keywords': ['contrataciones', 'contratacion', 'contratos', 'empresas'],
            'fetch': modelo_reporte.obtener_contrataciones_reporte,
            'headers': ['Empresa', 'RIF', 'N° Contrato', 'Monto', 'Descripción', 'Observación', 'Tipo', 'Modalidad', 'Objeto', 'Registro', 'Inicio', 'Adjudicación'],
            'fields': ['nombre_empresa', 'rif_empresa', 'numero_contrato', 'monto', 'descripcion', 'observacion', 'tipo_contrato', 'modalidad', 'objeto', 'fecha_registro', 'fecha_inicio', 'fecha_adjudicacion']
        },
        {
            'id': 'obras',
            'label': 'OBRAS',
            'keywords': ['obras', 'obra', 'proyectos', 'construccion', 'construcción'],
            'fetch': modelo_reporte.obtener_obras_reporte,
            'headers': ['Nombre de Obra', 'Ubicación', '% Avance', 'Semáforo', 'Color', 'Contratista'],
            'fields': ['nombre_obra', 'ubicacion_obra', 'porcentaje_avance_obra', 'semaforo', 'color', 'contratista']
        },
        {
            'id': 'publicaciones',
            'label': 'PUBLICACIONES',
            'keywords': ['publicaciones', 'publicacion', 'noticias', 'avisos', 'comunicados'],
            'fetch': modelo_reporte.obtener_publicaciones_reporte,
            'headers': ['Título', 'Autor', 'Fecha', 'Tipo'],
            'fields': ['titulo_publicacion', 'autor_publicacion', 'fecha_formateada', 'tipo_publicacion']
        }
    ]


def _detectar_tipo_filtro(filtros, modulo_id):
    if not filtros:
        return None
    if modulo_id == 'solicitudes' and any(k in filtros for k in ('cedula', 'correo', 'correo_dominio')):
        return 'solicitante'
    if modulo_id == 'obras' and any(k in filtros for k in ('ubicacion_obra', 'semaforo_estado', 'contratista', 'criticidad', 'nivel_gravedad', 'gerente')):
        return 'obra'
    return None


def _get_modulo_reconfigurado(mod_config, tipo_filtro):
    if tipo_filtro == 'solicitante' and mod_config['id'] == 'solicitudes':
        return {
            'id': mod_config['id'],
            'label': mod_config['label'],
            'keywords': mod_config['keywords'],
            'fetch': mod_config['fetch'],
            'headers': ['Cédula', 'Nombre del Solicitante', 'Correo', 'Teléfono', 'Descripción de la Solicitud', 'Fecha', 'Estatus', 'Municipio/Parroquia'],
            'fields': ['cedula', 'nombre_solicitante', 'correo', 'telefono', 'problematica', 'fecha', 'estatus_solicitud', 'municipio_parroquia']
        }
    if tipo_filtro == 'obra' and mod_config['id'] == 'obras':
        return {
            'id': mod_config['id'],
            'label': mod_config['label'],
            'keywords': mod_config['keywords'],
            'fetch': mod_config['fetch'],
            'headers': ['ID Obra', 'Nivel de Criticidad', 'Gerencia Responsable', 'Estatus', 'Avance (%)', 'Fecha de Registro'],
            'fields': ['id_obra', 'nivel_gravedad', 'gerente', 'semaforo', 'porcentaje_avance_obra', 'fecha_inicio']
        }
    return mod_config


def obtener_modulos_excel_por_filtro(filtro):
    modulos_config = construir_modulos_config_excel()
    if not filtro:
        return modulos_config
    texto = filtro.lower().strip()
    for mod in modulos_config:
        if texto == mod['id'].lower():
            return [mod]
    return []


CAMPOS_COMUNES = [
    'nombre_solicitante', 'cedula', 'telefono', 'correo', 'correo_dominio',
    'fecha_desde', 'fecha_hasta', 'municipio', 'direccion',
]

CAMPOS_POR_MODULO_EXCEL = {
    'solicitudes': [
        'tipo_solicitud', 'estatus_solicitud', 'problematica', 'cedula', 'nombre_solicitante',
        'fecha_desde', 'fecha_hasta', 'municipio', 'parroquia', 'direccion', 'telefono', 'correo',
        'correo_dominio', 'sector', 'ambito',
    ],
    'empleados': [
        'nombre_empleado', 'cargo', 'fecha_ingreso_desde', 'fecha_ingreso_hasta', 'estado_empleado',
        'cedula_persona', 'telefono', 'correo', 'direccion', 'parroquia', 'municipio', 'gerencia_asignada',
    ],
    'usuarios': [
        'nombre', 'cedula_usuario', 'correo', 'rol', 'estado',
    ],
    'contrataciones': [
        'empresa_ganadora', 'tipo_contrato', 'modalidad', 'objeto',
        'fecha_registro_desde', 'fecha_registro_hasta', 'fecha_inicio_procedimiento_desde',
        'fecha_inicio_procedimiento_hasta', 'fecha_adjudicacion_desde', 'fecha_adjudicacion_hasta',
        'numero_contrato',
    ],
    'obras': [
        'titulo_obra', 'ubicacion_obra', 'fecha_inicio_desde', 'fecha_inicio_hasta',
        'fecha_fin_desde', 'fecha_fin_hasta', 'semaforo_estado', 'contratista',
        'criticidad', 'nivel_gravedad', 'gerente',
    ],
    'publicaciones': [
        'titulo_publicacion', 'nombre_responsable', 'tipo_publicacion',
        'fecha_publicacion_desde', 'fecha_publicacion_hasta',
    ],
    'solicitantes': [
        'nombre', 'apellido', 'cedula', 'correo', 'correo_dominio',
    ],
}


def _colectar_filtros_form(modulo='general'):
    campos_especificos = CAMPOS_POR_MODULO_EXCEL.get(modulo, [])
    campos = list(dict.fromkeys(CAMPOS_COMUNES + campos_especificos))
    filtros = {}
    for campo in campos:
        val = request.form.get(campo, '').strip()
        if val:
            filtros[campo] = val
    return filtros


def _limpiar_filtros_por_modulo(filtros, modulo_id):
    campos_validos = {
        'solicitudes': {
            'tipo_solicitud', 'estatus_solicitud', 'problematica', 'cedula', 'nombre_solicitante',
            'fecha_desde', 'fecha_hasta', 'municipio', 'parroquia', 'direccion', 'telefono', 'correo',
            'correo_dominio', 'sector', 'ambito',
        },
        'empleados': {
            'nombre_empleado', 'cargo', 'fecha_ingreso_desde', 'fecha_ingreso_hasta', 'estado_empleado',
            'cedula_persona', 'telefono', 'correo', 'direccion', 'parroquia', 'municipio', 'gerencia_asignada',
        },
        'usuarios': {'nombre', 'cedula_usuario', 'correo', 'rol', 'estado'},
        'contrataciones': {
            'empresa_ganadora', 'tipo_contrato', 'modalidad', 'objeto',
            'fecha_registro_desde', 'fecha_registro_hasta', 'fecha_inicio_procedimiento_desde',
            'fecha_inicio_procedimiento_hasta', 'fecha_adjudicacion_desde', 'fecha_adjudicacion_hasta',
            'numero_contrato',
        },
        'obras': {
            'titulo_obra', 'ubicacion_obra', 'fecha_inicio_desde', 'fecha_inicio_hasta',
            'fecha_fin_desde', 'fecha_fin_hasta', 'semaforo_estado', 'contratista',
            'criticidad', 'nivel_gravedad', 'gerente',
        },
        'publicaciones': {
            'titulo_publicacion', 'nombre_responsable', 'tipo_publicacion',
            'fecha_publicacion_desde', 'fecha_publicacion_hasta',
        },
        'solicitantes': {'nombre', 'apellido', 'cedula', 'correo', 'correo_dominio'},
    }
    validos = campos_validos.get(modulo_id, set())
    return {k: v for k, v in filtros.items() if k in validos}


@reporte_excel_bp.route('/reporte-excel', methods=['GET', 'POST'])
def generarReporteExcel():
    if 'conectado' not in session:
        flash('Primero debes iniciar sesión.', 'error')
        return redirect(url_for('login_bp.inicio'))

    if request.method == 'POST':
        modulo = request.form.get('modulo', '').strip().lower()
        filtros = _colectar_filtros_form(modulo)

        modulos_config = construir_modulos_config_excel()
        if modulo and modulo != 'general':
            modulos_a_procesar = [m for m in modulos_config if m['id'] == modulo]
        else:
            modulos_a_procesar = modulos_config

        if not modulos_a_procesar:
            flash('No se encontró ningún módulo relacionado con el criterio ingresado.', 'info')
            return redirect(url_for('reporte_excel_bp.generarReporteExcel'))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for mod in modulos_a_procesar:
            mod_filtros = _limpiar_filtros_por_modulo(filtros, mod['id'])
            tipo_filtro = _detectar_tipo_filtro(mod_filtros, mod['id'])
            mod_config = _get_modulo_reconfigurado(mod, tipo_filtro)
            data = mod_config['fetch'](mod_filtros if mod_filtros else None)
            if not data:
                continue

            sheet_title = re.sub(r'[\\\/\?\*\[\]]', ' ', mod_config['label'])
            ws = wb.create_sheet(title=sheet_title)
            ws.sheet_view.showGridLines = True

            if os.path.exists(LOGO_GOBERNACION):
                try:
                    img_izq = XLImage(LOGO_GOBERNACION)
                    img_izq.width = 140
                    img_izq.height = 42
                    ws.add_image(img_izq, 'A2')
                except Exception:
                    pass
            if os.path.exists(LOGO_INVILARA):
                try:
                    img_der = XLImage(LOGO_INVILARA)
                    img_der.width = 130
                    img_der.height = 40
                    total_cols = len(mod_config['headers'])
                    right_col_letter = get_column_letter(total_cols)
                    ws.add_image(img_der, f'{right_col_letter}2')
                except Exception:
                    pass

            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 48
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 16
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 20

            total_cols = len(mod_config['headers'])
            last_col_letter = get_column_letter(total_cols)
            fecha_cell = ws.cell(row=1, column=total_cols, value=f'Fecha de Emisión: {datetime.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")}')
            fecha_cell.font = Font(italic=True, size=8, color='666666')
            fecha_cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)

            header_title = ws.cell(row=3, column=1, value='GOBERNACIÓN DEL ESTADO LARA - INVILARA')
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
            header_title.font = Font(bold=True, size=11)
            header_title.alignment = Alignment(horizontal='center', vertical='center')

            header_subtitle = ws.cell(row=4, column=1, value='REPORTE DETALLADO DE GESTIÓN')
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
            header_subtitle.font = Font(bold=True, size=9, color='666666')
            header_subtitle.alignment = Alignment(horizontal='center', vertical='center')

            start_row = 6
            ws.cell(row=start_row, column=1, value=sheet_title)
            ws.row_dimensions[start_row].height = 24
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(mod_config['headers']))
            title_cell = ws.cell(row=start_row, column=1)
            title_cell.font = Font(bold=True, size=13, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1

            column_widths = {i: len(str(h)) for i, h in enumerate(mod_config['headers'], 1)}

            for col_num, header in enumerate(mod_config['headers'], 1):
                cell = ws.cell(row=start_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = header_alignment
                column_widths[col_num] = max(column_widths[col_num], len(str(header)))
            start_row += 1

            date_fields = {'fecha', 'fecha_ingreso', 'fecha_formateada', 'fecha_registro', 'fecha_inicio', 'fecha_adjudicacion'}

            for reg in data:
                for col_num, field in enumerate(mod_config['fields'], 1):
                    val = reg.get(field, '')
                    if field in date_fields and val:
                        val = str(val)
                    elif isinstance(val, datetime.datetime):
                        val = val.strftime('%d/%m/%Y %H:%M')
                    elif isinstance(val, datetime.date):
                        val = val.strftime('%d/%m/%Y')
                    else:
                        val = str(val) if val is not None else ''

                    cell = ws.cell(row=start_row, column=col_num, value=val)
                    cell.border = border
                    cell.alignment = cell_alignment
                    column_widths[col_num] = max(column_widths[col_num], min(len(str(val)), 60))
                start_row += 1

            for col_num, width in column_widths.items():
                adjusted = min(max(width + 3, 12), 50)
                ws.column_dimensions[get_column_letter(col_num)].width = adjusted

        if not wb.sheetnames:
            flash('No hay registros disponibles para generar el reporte con los criterios ingresados.', 'info')
            return redirect(url_for('reporte_excel_bp.generarReporteExcel'))

        filename = f"Reporte_Invilara_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        base_dir = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.normpath(os.path.join(base_dir, '..', 'static', 'downloads-excel'))
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        ruta_archivo = os.path.join(folder_path, filename)
        wb.save(ruta_archivo)
        BitacoraService.registrar_accion(
            session, 'Reportes', 'VER',
            f'Generó un reporte Excel de gestión'
        )
        return send_file(ruta_archivo, as_attachment=True)

    publicaciones_previa = modelo_reporte.obtener_publicaciones_reporte()
    return render_template('reportes/reporteExcel.html', publicaciones=publicaciones_previa)
